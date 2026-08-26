from __future__ import annotations
from dataclasses import dataclass, asdict
from pathlib import Path
from datetime import datetime, timezone
import hashlib, json
import pandas as pd
from openpyxl import Workbook

@dataclass(frozen=True)
class LedgerEvent:
    event_id: str
    event_type: str
    created_at: str
    aggregate_id: str
    payload: dict
    prev_hash: str
    event_hash: str

class ImmutableEventLedger:
    def __init__(self, path='artifacts/paper_trading/v21_event_ledger.jsonl', unit_brl=500.0):
        self.path=Path(path); self.path.parent.mkdir(parents=True,exist_ok=True); self.unit_brl=float(unit_brl)
    def _rows(self): return [json.loads(x) for x in self.path.read_text(encoding='utf-8').splitlines() if x.strip()] if self.path.exists() else []
    def _next_hash(self):
        rows=self._rows(); return rows[-1]['event_hash'] if rows else 'GENESIS'
    def append(self,event_type,aggregate_id,payload,event_id=None):
        event_id=event_id or hashlib.sha256(f'{event_type}|{aggregate_id}|{datetime.now(timezone.utc).isoformat()}|{json.dumps(payload,sort_keys=True,default=str)}'.encode()).hexdigest()[:32]
        created=datetime.now(timezone.utc).isoformat(); prev=self._next_hash()
        raw=f'{event_id}|{event_type}|{created}|{aggregate_id}|{prev}|{json.dumps(payload,sort_keys=True,default=str)}'
        event_hash=hashlib.sha256(raw.encode()).hexdigest()
        if any(r['event_id']==event_id for r in self._rows()): raise ValueError('DUPLICATE_EVENT_ID')
        row=asdict(LedgerEvent(event_id,event_type,created,aggregate_id,payload,prev,event_hash))
        with self.path.open('a',encoding='utf-8') as f:f.write(json.dumps(row,sort_keys=True,separators=(',',':'))+'\n')
        return row
    def events(self): return self._rows()
    def verify_chain(self):
        prev='GENESIS'
        for r in self._rows():
            if r['prev_hash']!=prev:return {'valid':False,'reason':'BROKEN_CHAIN','event_id':r['event_id']}
            raw=f"{r['event_id']}|{r['event_type']}|{r['created_at']}|{r['aggregate_id']}|{r['prev_hash']}|{json.dumps(r['payload'],sort_keys=True,default=str)}"
            if hashlib.sha256(raw.encode()).hexdigest()!=r['event_hash']:return {'valid':False,'reason':'HASH_MISMATCH','event_id':r['event_id']}
            prev=r['event_hash']
        return {'valid':True,'events':len(self._rows()),'fingerprint':self.fingerprint()}
    def fingerprint(self): return hashlib.sha256(self.path.read_bytes()).hexdigest() if self.path.exists() else None
    def decisions(self):
        return [r['payload'] for r in self._rows() if r['event_type'] in ('SIGNAL_CREATED','SIGNAL_REJECTED')]
    def positions(self):
        rows={}
        for e in self._rows():
            if e['event_type']=='SIGNAL_CREATED' and e['payload'].get('decision')=='BET': rows[e['aggregate_id']]={**dict(e['payload']),'aggregate_id':e['aggregate_id']}
            elif e['event_type'] in ('POSITION_UPDATED','POSITION_EXITED','RESULT_SETTLED') and e['aggregate_id'] in rows:
                rows[e['aggregate_id']].update(e['payload'])
        return list(rows.values())
    def performance(self):
        rows=self.positions(); settled=[r for r in rows if r.get('status')=='SETTLED']; stake=sum(float(r.get('stake_units',0)) for r in settled); pnl=sum(float(r.get('pnl_units',0)) for r in settled)
        return {'positions':len(rows),'settled':len(settled),'wins':sum(r.get('result')=='WIN' for r in settled),'losses':sum(r.get('result')=='LOSS' for r in settled),'units':pnl,'brl':pnl*self.unit_brl,'roi':pnl/stake if stake else None}
    def export_xlsx(self,path='artifacts/paper_trading/v21_results.xlsx'):
        p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);wb=Workbook();ws=wb.active;ws.title='DASHBOARD';perf=self.performance()
        for row in [['Metric','Value'],*[[k,v] for k,v in perf.items()]]:ws.append(row)
        rows=self.positions();cols=['created_at','event_id','league','event_name','market','selection','odds','stake_units','stake_brl','decision','result','pnl_units','fair_probability','fair_odds','edge','ev','clv','mode','action','why']
        sh=wb.create_sheet('RESULTADOS');sh.append(cols)
        for r in rows:sh.append([r.get(c) for c in cols])
        for name,key in [('MERCADOS','market'),('LIGAS','league')]:
            sh=wb.create_sheet(name);sh.append([key,'positions','settled','units','ROI'])
            d=pd.DataFrame(rows)
            if not d.empty:
                for value,g in d.groupby(key,dropna=False):
                    ss=g[g.get('status','')=='SETTLED'] if 'status' in g else g.iloc[0:0]
                    stake=float(pd.to_numeric(ss.get('stake_units',pd.Series(dtype=float)),errors='coerce').fillna(0).sum()) if not ss.empty else 0
                    pnl=float(pd.to_numeric(ss.get('pnl_units',pd.Series(dtype=float)),errors='coerce').fillna(0).sum()) if not ss.empty else 0
                    sh.append([value,len(g),len(ss),pnl,pnl/stake if stake else None])
        nb=[r for r in self.decisions() if r.get('decision')=='NO BET'];sh=wb.create_sheet('NO_BET');sh.append(cols)
        for r in nb:sh.append([r.get(c) for c in cols])
        wb.save(p);return str(p)
