from dataclasses import dataclass,asdict
from pathlib import Path
from datetime import datetime,timezone
import hashlib,json
import pandas as pd
from openpyxl import Workbook
@dataclass(frozen=True)
class LedgerRecord:
    record_id:str; event_id:str; timestamp:str; league:str; country:str; season:str; market:str; selection:str; odds:float; stake_units:float; stake_brl:float; result:str|None; pnl_units:float; fair_probability:float; fair_odds:float|None; edge:float; ev:float; clv:float|None; model_version:str; feature_snapshot:str; pit_status:str; decision:str; entry_minute:int|None=None; scoreline:str|None=None; exit_reason:str|None=None; exit_timestamp:str|None=None; status:str='OPEN'
class LedgerV20:
    def __init__(self,path='artifacts/paper_trading/v20_ledger.jsonl',unit_brl=500.0): self.path=Path(path);self.path.parent.mkdir(parents=True,exist_ok=True);self.unit_brl=unit_brl
    def _read(self): return [json.loads(x) for x in self.path.read_text(encoding='utf-8').splitlines() if x.strip()] if self.path.exists() else []
    def append(self,record):
        rows=self._read()
        if any(r['record_id']==record.record_id for r in rows): raise ValueError('LEDGER_IMMUTABLE_DUPLICATE_ID')
        with self.path.open('a',encoding='utf-8') as f:f.write(json.dumps(asdict(record),sort_keys=True,separators=(',',':'))+'\n')
        return asdict(record)
    def rows(self): return self._read()
    def settle(self,record_id,result,closing_odds=None,exit_reason=None):
        rows=self._read();target=next((r for r in rows if r['record_id']==record_id),None)
        if not target: raise KeyError(record_id)
        if target['status']!='OPEN': raise ValueError('LEDGER_SETTLEMENT_ALREADY_FINAL')
        if result not in {'WIN','LOSS','VOID'}: raise ValueError('INVALID_RESULT')
        pnl=target['stake_units']*(target['odds']-1) if result=='WIN' else -target['stake_units'] if result=='LOSS' else 0.0
        clv=(target['odds']/float(closing_odds)-1) if closing_odds and float(closing_odds)>1 else None
        target.update({'result':result,'pnl_units':pnl,'clv':clv,'status':'SETTLED','exit_reason':exit_reason,'exit_timestamp':datetime.now(timezone.utc).isoformat()})
        tmp=self.path.with_suffix('.tmp');tmp.write_text(''.join(json.dumps(r,sort_keys=True,separators=(',',':'))+'\n' for r in rows),encoding='utf-8');tmp.replace(self.path);return target
    def export_xlsx(self,path='artifacts/paper_trading/v20_results.xlsx'):
        rows=self._read();df=pd.DataFrame(rows);p=Path(path);p.parent.mkdir(parents=True,exist_ok=True);wb=Workbook();ws=wb.active;ws.title='DASHBOARD';settled=df[df.status=='SETTLED'] if not df.empty else df;stake=float(settled.stake_units.sum()) if not settled.empty else 0.;pnl=float(settled.pnl_units.sum()) if not settled.empty else 0.
        for r in [['Metric','Value'],['Bets',len(df)],['Settled',len(settled)],['Units',pnl],['ROI',pnl/stake if stake else None],['BRL',pnl*self.unit_brl]]:ws.append(r)
        cols=['date','event_id','league','market','selection','odds','stake_units','stake_brl','result','pnl_units','fair_probability','fair_odds','edge','ev','clv']
        def add_sheet(name,frame):
            sh=wb.create_sheet(name);sh.append(cols)
            for _,r in frame.iterrows():sh.append([str(r.get('timestamp',''))[:10],r.get('event_id'),r.get('league'),r.get('market'),r.get('selection'),r.get('odds'),r.get('stake_units'),r.get('stake_brl'),r.get('result'),r.get('pnl_units'),r.get('fair_probability'),r.get('fair_odds'),r.get('edge'),r.get('ev'),r.get('clv')])
        add_sheet('RESULTADOS',df)
        for name,col in [('MERCADOS','market'),('LIGAS','league')]:
            sh=wb.create_sheet(name);sh.append([col,'bets','units','ROI'])
            if not df.empty and not settled.empty:
                g=settled.groupby(col).agg(bets=('record_id','count'),units=('pnl_units','sum'),stake=('stake_units','sum')).reset_index();g['ROI']=g['units']/g['stake'].replace(0,pd.NA)
                for x in g.itertuples(index=False,name=None):sh.append(list(x))
        if not df.empty:
            add_sheet('NO_BET',df[df.decision=='NO BET']);add_sheet('LIVE',df[df.entry_minute.astype('float64').fillna(0)>0])
        wb.save(p);return str(p)
    def fingerprint(self): return hashlib.sha256(self.path.read_bytes()).hexdigest() if self.path.exists() else None
