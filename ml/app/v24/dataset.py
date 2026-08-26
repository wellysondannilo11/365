from __future__ import annotations
from pathlib import Path
from datetime import datetime, timezone
import json, hashlib, math
import pandas as pd
from openpyxl import Workbook
from .hashchain import HashChain

class EmpiricalDatasetV24:
    def __init__(self,path="data/research/robo_bet_dataset_v24.jsonl"):
        self.path=Path(path); self.chain=HashChain(self.path)
    def append(self,record):
        row=dict(record); row.setdefault("dataset_version","v24"); row.setdefault("created_at",datetime.now(timezone.utc).isoformat())
        row.setdefault("observation_id",hashlib.sha256(f"{row.get('event_id')}|{row.get('snapshot_id')}|{row.get('decision_id')}|{row.get('created_at')}".encode()).hexdigest()[:32])
        if row.get("mode") not in {"PAPER","SHADOW"}: raise ValueError("REAL_MONEY_MODE_FORBIDDEN")
        return self.chain.append(row)
    def rows(self): return self.chain.rows()
    def verify(self): return self.chain.verify()
    def stats(self):
        rows=self.rows()
        return {"rows":len(rows),"events":len({r.get("event_id") for r in rows if r.get("event_id")}),
                "snapshots":len({r.get("snapshot_id") for r in rows if r.get("snapshot_id")}),
                "decisions":sum(1 for r in rows if r.get("decision")),
                "bets":sum(r.get("decision")=="BET" for r in rows),
                "no_bets":sum(r.get("decision")=="NO BET" for r in rows),
                "paper":sum(r.get("mode")=="PAPER" for r in rows),
                "shadow":sum(r.get("mode")=="SHADOW" for r in rows),
                "hash_chain":self.verify()}
    def performance(self, mode=None):
        rows=[r for r in self.rows() if (mode is None or r.get("mode")==mode) and r.get("result") in {"WIN","LOSS","VOID"}]
        stake=sum(float(r.get("stake_units") or 0) for r in rows)
        pnl=sum(float(r.get("pnl_units") or 0) for r in rows)
        return {"mode":mode or "ALL","settled":len(rows),"wins":sum(r.get("result")=="WIN" for r in rows),
                "losses":sum(r.get("result")=="LOSS" for r in rows),"pnl_units":pnl,"roi":pnl/stake if stake else None,
                "stake_units":stake}
    def breakdown(self,key):
        rows=self.rows(); out={}
        for r in rows:
            if r.get("result") not in {"WIN","LOSS","VOID"}: continue
            k=str(r.get(key) or "UNKNOWN"); x=out.setdefault(k,{"bets":0,"wins":0,"stake_units":0.0,"pnl_units":0.0})
            x["bets"]+=1;x["wins"]+=r.get("result")=="WIN";x["stake_units"]+=float(r.get("stake_units") or 0);x["pnl_units"]+=float(r.get("pnl_units") or 0)
        for x in out.values(): x["roi"]=x["pnl_units"]/x["stake_units"] if x["stake_units"] else None
        return out

    def export_xlsx(self,path="artifacts/paper_trading/v24_results.xlsx"):
        from pathlib import Path
        p=Path(path);p.parent.mkdir(parents=True,exist_ok=True)
        rows=self.rows(); wb=Workbook(); ws=wb.active; ws.title="DASHBOARD"
        ws.append(["Metric","Value"])
        for k,v in self.performance().items(): ws.append([k,v])
        for mode in ("PAPER","SHADOW"):
            sh=wb.create_sheet(mode); cols=["created_at","decision_time","event_name","league","market","selection","bookmaker","odds","fair_odds","edge","ev","stake_units","decision","result","pnl_units","clv","model_version","mode"]
            sh.append(cols)
            for r in rows:
                if r.get("mode")==mode: sh.append([r.get(c) for c in cols])
        for sheet,key in (("MERCADOS","market"),("LIGAS","league")):
            sh=wb.create_sheet(sheet); sh.append([key,"bets","wins","stake_units","pnl_units","ROI"])
            for name,x in self.breakdown(key).items(): sh.append([name,x["bets"],x["wins"],x["stake_units"],x["pnl_units"],x["roi"]])
        sh=wb.create_sheet("NO_BET"); sh.append(["decision_time","event_name","league","market","selection","odds","fair_odds","edge","ev","reason","mode"])
        for r in rows:
            if r.get("decision")=="NO BET": sh.append([r.get(c) for c in ["decision_time","event_name","league","market","selection","odds","fair_odds","edge","ev","reason","mode"]])
        wb.save(p); return str(p)
