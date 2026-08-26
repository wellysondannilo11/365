from __future__ import annotations
from pathlib import Path
from datetime import datetime, timezone
import hashlib,json,statistics
from openpyxl import Workbook

class V25Dataset:
    def __init__(self,path='data/research/robo_bet_dataset_v25.jsonl',persistence=None):
        self.path=Path(path); self.path.parent.mkdir(parents=True,exist_ok=True); self.persistence=persistence
    def _local_rows(self):
        if not self.path.exists(): return []
        return [json.loads(x) for x in self.path.read_text(encoding='utf-8').splitlines() if x.strip()]
    def _rows(self):
        if self.persistence is not None and self.persistence.available:
            rows=self.persistence.rows()
            if rows is not None:return rows
        return self._local_rows()
    def _stable_observation_id(self,row):
        identity={k:row.get(k) for k in ('event_id','snapshot_id','decision','mode','market','selection','line','odds','bookmaker')}
        return hashlib.sha256(json.dumps(identity,sort_keys=True,separators=(',',':'),default=str).encode()).hexdigest()[:32]
    def append(self,row):
        row=dict(row); mode=str(row.get('mode','SHADOW')).upper()
        if mode not in {'PAPER','SHADOW'}: raise ValueError('REAL_MONEY_MODE_FORBIDDEN')
        row.setdefault('dataset_version','v25'); row.setdefault('created_at',datetime.now(timezone.utc).isoformat())
        row.setdefault('observation_id',self._stable_observation_id(row))
        existing=self._rows()
        duplicate=next((x for x in existing if x.get('observation_id')==row['observation_id']),None)
        if duplicate:return duplicate
        row['previous_hash']=existing[-1].get('row_hash') if existing else (self.persistence.head_hash() if self.persistence and self.persistence.available else None)
        canonical=json.dumps(row,sort_keys=True,separators=(',',':'),default=str)
        row['row_hash']=hashlib.sha256(canonical.encode()).hexdigest()
        if self.persistence is not None and self.persistence.available:
            stored=self.persistence.append_row(row)
            if stored is None: raise RuntimeError('PERSISTENCE_WRITE_FAILED')
            row=stored
        # Always keep a local forensic mirror. It is not the primary source when DB is online.
        with self.path.open('a',encoding='utf-8') as f:f.write(json.dumps(row,sort_keys=True,separators=(',',':'),default=str)+'\n')
        return row
    def rows(self):return self._rows()
    def verify(self):
        rows=self._rows();prev=None
        for i,r in enumerate(rows):
            if r.get('previous_hash')!=prev:return {'valid':False,'index':i,'reason':'BROKEN_PREVIOUS_HASH'}
            x=dict(r); expected=x.pop('row_hash',None)
            if hashlib.sha256(json.dumps(x,sort_keys=True,separators=(',',':'),default=str).encode()).hexdigest()!=expected:return {'valid':False,'index':i,'reason':'ROW_HASH_MISMATCH'}
            prev=expected
        return {'valid':True,'rows':len(rows),'head':prev,'source':'POSTGRESQL' if self.persistence is not None and self.persistence.available else 'JSONL_FALLBACK'}
    def stats(self):
        r=self.rows(); decisions=[x for x in r if x.get('decision')]
        return {'rows':len(r),'events':len({x.get('event_id') for x in r if x.get('event_id')}),'snapshots':len({x.get('snapshot_id') for x in r if x.get('snapshot_id')}),'decisions':len(decisions),'bets':sum(x.get('decision')=='BET' for x in r),'no_bets':sum(x.get('decision')=='NO BET' for x in r),'paper':sum(x.get('mode')=='PAPER' for x in r),'shadow':sum(x.get('mode')=='SHADOW' for x in r),'hash_chain':self.verify()}
    def performance(self,mode=None):
        r=[x for x in self.rows() if (mode is None or x.get('mode')==mode) and x.get('result') in {'WIN','LOSS','VOID','HALF_WIN','PUSH','HALF_LOSS'}]
        stake=sum(float(x.get('stake_units') or 0) for x in r); pnl=sum(float(x.get('pnl_units') or 0) for x in r)
        return {'mode':mode or 'ALL','settled':len(r),'wins':sum(x.get('result')=='WIN' for x in r),'losses':sum(x.get('result')=='LOSS' for x in r),'stake_units':stake,'pnl_units':pnl,'roi':pnl/stake if stake else None}
    def breakdown(self,key):
        out={}
        for r in self.rows():
            if r.get('result') not in {'WIN','LOSS','VOID','HALF_WIN','PUSH','HALF_LOSS'}:continue
            k=str(r.get(key) or 'UNKNOWN'); x=out.setdefault(k,{'bets':0,'wins':0,'stake_units':0.0,'pnl_units':0.0,'clv':[]}); x['bets']+=1; x['wins']+=r.get('result')=='WIN'; x['stake_units']+=float(r.get('stake_units') or 0); x['pnl_units']+=float(r.get('pnl_units') or 0)
            if r.get('clv') is not None:x['clv'].append(float(r['clv']))
        for x in out.values():x['roi']=x['pnl_units']/x['stake_units'] if x['stake_units'] else None;x['avg_clv']=statistics.mean(x['clv']) if x['clv'] else None;x.pop('clv',None)
        return out
    def export_xlsx(self,path='artifacts/paper_trading/v25_results.xlsx'):
        p=Path(path);p.parent.mkdir(parents=True,exist_ok=True); rows=self.rows(); wb=Workbook(); dash=wb.active;dash.title='DASHBOARD';dash.append(['Metric','Value'])
        for k,v in {**self.stats(),**{'pnl_units':self.performance()['pnl_units'],'roi':self.performance()['roi']}}.items():dash.append([k,str(v)])
        # Operational periods are computed from settled real PAPER/SHADOW rows only.
        from datetime import datetime, timezone
        now=datetime.now(timezone.utc); day=now.date().isoformat(); month=now.strftime('%Y-%m'); year=now.strftime('%Y')
        for title,prefix in (('HOJE',day),('MES',month),('ANO',year)):
            sh=wb.create_sheet(title); sh.append(['Metric','Value'])
            subset=[r for r in rows if r.get('result') and str(r.get('created_at','')).startswith(prefix)]
            stake=sum(float(r.get('stake_units') or 0) for r in subset); pnl=sum(float(r.get('pnl_units') or 0) for r in subset)
            sh.append(['Apostas',len(subset)]); sh.append(['Unidades',pnl]); sh.append(['ROI',pnl/stake if stake else None]); sh.append(['Win rate',sum(r.get('result')=='WIN' for r in subset)/len(subset) if subset else None]); sh.append(['CLV',statistics.mean([float(r['clv']) for r in subset if r.get('clv') is not None]) if any(r.get('clv') is not None for r in subset) else None])
        cols=['created_at','decision_time','event_name','league','market','line','selection','bookmaker','odds','closing_odds','fair_probability','fair_odds','edge','ev','stake_units','result','pnl_units','clv','phase','mode','decision','decision_id','observation_id','snapshot_id','source_timestamp','captured_at']
        for name,mode in (('PAPER','PAPER'),('SHADOW','SHADOW')):
            sh=wb.create_sheet(name);sh.append(cols);[sh.append([r.get(c) for c in cols]) for r in rows if r.get('mode')==mode]
        for sheet,key in (('RESULTADOS',None),('MERCADOS','market'),('LIGAS','league'),('BOOKMAKERS','bookmaker')):
            sh=wb.create_sheet(sheet)
            if key: sh.append([key,'bets','wins','stake_units','pnl_units','ROI','avg_clv']); data=self.breakdown(key)
            else: sh.append(cols); data=None
            if data is not None:
                for k,v in data.items():sh.append([k,v['bets'],v['wins'],v['stake_units'],v['pnl_units'],v['roi'],v['avg_clv']])
            else:[sh.append([r.get(c) for c in cols]) for r in rows if r.get('result')]
        sh=wb.create_sheet('NO BET');sh.append(['decision_time','event_name','league','market','line','selection','bookmaker','odds','fair_odds','edge','ev','reason','mode','decision_id'])
        [sh.append([r.get(c) for c in ['decision_time','event_name','league','market','line','selection','bookmaker','odds','fair_odds','edge','ev','reason','mode','decision_id']]) for r in rows if r.get('decision')=='NO BET']
        sh=wb.create_sheet('PRICE MOVEMENT');sh.append(['event_id','bookmaker','market','selection','line','opening_price','current_price','movement','velocity','acceleration'])
        for r in rows:
            pm=r.get('price_movement') or {};sh.append([r.get('event_id'),r.get('bookmaker'),r.get('market'),r.get('selection'),r.get('line'),pm.get('opening_price'),pm.get('current_price'),pm.get('movement'),pm.get('velocity'),pm.get('acceleration')])
        sh=wb.create_sheet('POSITIONS');sh.append(['position_id','event_id','decision_id','status','entry_odds','stake_units','opened_at','closed_at','action'])
        for r in rows:
            if r.get('position_id'):sh.append([r.get('position_id'),r.get('event_id'),r.get('decision_id'),r.get('position_state'),r.get('odds'),r.get('stake_units'),r.get('opened_at'),r.get('closed_at'),r.get('decision')])
        sh=wb.create_sheet('CARD MARKETS'); sh.append(['created_at','decision_time','event_id','event_name','league','market','selection','line','bookmaker','odds','fair_probability','fair_odds','edge','ev','stake_units','result','pnl_units','clv','mode','decision','reason','card_model','card_model_version','card_feature_version','referee_id','referee_cards_avg','team_cards_avg','opponent_cards_avg','h2h_cards_avg','match_importance','match_intensity','cards_observed','cards_remaining_expected','snapshot_id','observation_id','source_timestamp','captured_at'])
        for r in rows:
            if str(r.get('market','')).upper() in {'CARD_TOTALS','CARD_HOME','CARD_AWAY'}:
                sh.append([r.get(c) for c in ['created_at','decision_time','event_id','event_name','league','market','selection','line','bookmaker','odds','fair_probability','fair_odds','edge','ev','stake_units','result','pnl_units','clv','mode','decision','reason','card_model','card_model_version','card_feature_version','referee_id','referee_cards_avg','team_cards_avg','opponent_cards_avg','h2h_cards_avg','match_importance','match_intensity','cards_observed','cards_remaining_expected','snapshot_id','observation_id','source_timestamp','captured_at']])
        sh=wb.create_sheet('CARD NO BET'); sh.append(['decision_time','event_id','market','selection','line','bookmaker','odds','fair_odds','edge','ev','reason','mode','decision_id'])
        for r in rows:
            if str(r.get('market','')).upper() in {'CARD_TOTALS','CARD_HOME','CARD_AWAY'} and r.get('decision')=='NO BET':
                sh.append([r.get(c) for c in ['decision_time','event_id','market','selection','line','bookmaker','odds','fair_odds','edge','ev','reason','mode','decision_id']])
        sh=wb.create_sheet('CARD PERFORMANCE'); sh.append(['market','bets','wins','stake_units','pnl_units','ROI','avg_clv'])
        for k,v in self.breakdown('market').items():
            if str(k).upper() in {'CARD_TOTALS','CARD_HOME','CARD_AWAY'}: sh.append([k,v['bets'],v['wins'],v['stake_units'],v['pnl_units'],v['roi'],v['avg_clv']])
        sh=wb.create_sheet('CARD PRICE MOVEMENT'); sh.append(['event_id','bookmaker','market','selection','line','opening_price','current_price','movement','velocity','acceleration'])
        for r in rows:
            if str(r.get('market','')).upper() in {'CARD_TOTALS','CARD_HOME','CARD_AWAY'}:
                pm=r.get('price_movement') or {}; sh.append([r.get('event_id'),r.get('bookmaker'),r.get('market'),r.get('selection'),r.get('line'),pm.get('opening_price'),pm.get('current_price'),pm.get('movement'),pm.get('velocity'),pm.get('acceleration')])
        sh=wb.create_sheet('CARD LIVE'); sh.append(['event_id','decision_time','minute','cards_observed','cards_remaining_expected','final_expected_cards','live_fair_probability','live_fair_odds','decision'])
        for r in rows:
            if str(r.get('market','')).upper() in {'CARD_TOTALS','CARD_HOME','CARD_AWAY'} and r.get('phase')=='LIVE':
                sh.append([r.get('event_id'),r.get('decision_time'),r.get('minute'),r.get('cards_observed'),r.get('cards_remaining_expected'),r.get('final_expected_cards'),r.get('fair_probability'),r.get('fair_odds'),r.get('decision')])
        sh=wb.create_sheet('CARD REFEREES'); sh.append(['referee_id','referee_cards_avg','sample_size','source','source_timestamp','captured_at','quality','feature_version'])
        seen_ref=set()
        for r in rows:
            rid=r.get('referee_id')
            if rid and rid not in seen_ref:
                seen_ref.add(rid); sh.append([rid,r.get('referee_cards_avg'),r.get('referee_sample_size'),r.get('source'),r.get('source_timestamp'),r.get('captured_at'),r.get('data_quality'),r.get('card_feature_version')])
        sh=wb.create_sheet('SNAPSHOTS');sh.append(['snapshot_id','event_id','event_name','league','phase','bookmaker','market','selection','line','odds','source_timestamp','captured_at','received_at','quality_status','feed_age_seconds'])
        snapshot_path=Path('data/research/robo_bet_snapshots_v25.jsonl')
        if snapshot_path.exists():
            for line in snapshot_path.read_text(encoding='utf-8').splitlines():
                if not line.strip(): continue
                r=json.loads(line); sh.append([r.get(c) for c in ['snapshot_id','event_id','event_name','league','phase','bookmaker','market','selection','line','odds','source_timestamp','captured_at','received_at','quality_status','feed_age_seconds']])
        wb.save(p);return str(p)
